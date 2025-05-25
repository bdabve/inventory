#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# Functions to work with for all project

import os
from datetime import date, datetime
from ummalqura.hijri_date import HijriDate

from django.contrib import messages
from django.shortcuts import redirect
from django.utils.timezone import is_aware
from django.db import transaction
from django.db.models import Q, F
from . import models
from . import forms

import openpyxl
from openpyxl.utils import get_column_letter


def filter_articles(search_word=None, category_slug=None, stock_alarm=False, art_sans_prix=False):
    """
    Filters articles based on search word, category slug, stock alarm, and art_sans_prix.
    This work with article list view
    """
    articles = models.Article.objects.select_related('category').all()

    if stock_alarm:
        articles = articles.filter(qte=0)

    if art_sans_prix:
        articles = articles.filter(prix=0)

    if category_slug:
        try:
            category = models.Category.objects.get(slug=category_slug)
            articles = articles.filter(category=category)
        except models.Category.DoesNotExist:
            category = None
    else:
        category = None

    if search_word:
        articles = articles.filter(
            Q(designation__icontains=search_word) |
            Q(code__icontains=search_word) |
            Q(ref__icontains=search_word)
        ).distinct()

    return articles, category, len(articles)


def process_stock_entry(article, new_qte, new_prix, entree_date, user):
    """
    Process stock entry update with quantity and price adjustment.
    This function creates a new Movement record and updates the Article's stock and price.

    Parameters:
    - article: Article instance
    - new_qte: quantity to add
    - new_prix: price of the incoming stock
    - entree_date: date of entry
    - user: User performing the entry
    """
    try:
        with transaction.atomic():
            # Save movement
            models.Movement.objects.create(
                art_id=article,
                movement_date=entree_date,
                user_id=user,
                movement="Entree",
                qte=new_qte,
                prix=new_prix,
            )

            # Update article stock and price
            if article.prix == new_prix:
                article.qte = F('qte') + new_qte
            else:
                total_val = (new_qte * new_prix) + (article.valeur or 0)
                total_qte = article.qte + new_qte
                article.prix = total_val / total_qte
                article.qte = F('qte') + new_qte

            article.save()
            return True
    except Exception as err:
        return f"Erreur lors de l'entrée: {str(err)}"


def process_stock_sortie(article, sortie_qte, sortie_date, user):
    """
    Process stock Sortie: updates article quantity and logs the movement.

    Parameters:
    - article: Article instance
    - sortie_qte: quantity to subtract
    - sortie_date: date of exit
    - user: User performing the sortie
    """
    try:
        with transaction.atomic():
            # Log the sortie movement
            models.Movement.objects.create(
                art_id=article,
                user_id=user,
                movement_date=sortie_date,
                movement="Sortie",
                qte=sortie_qte,
                prix=article.prix,
            )

            # Update the quantity
            article.qte = F('qte') - sortie_qte
            article.save()
        return True
    except Exception as e:
        return f"Erreur lors de la sortie: {str(e)}"


# == Article Details Functions
def detail_search_article(request, current_art_id, current_slug):
    """Search Article from the Details Page"""
    form = forms.SearchArticleForm(request.POST)
    if form.is_valid():
        code = form.cleaned_data['search_word'].upper()
        try:
            art = models.Article.objects.get(code=code)
            return redirect('magasin:article_detail', art_id=art.art_id, slug=art.slug)
        except models.Article.DoesNotExist:
            messages.warning(request, f'Aucun article trouvé avec le code « {code} »')
            return redirect('magasin:article_detail', art_id=current_art_id, slug=current_slug)
    return None


def detail_handle_entree(request, article):
    form = forms.EntreeForm(request.POST)
    if form.is_valid():
        new_qte = form.cleaned_data['qte']
        new_prix = form.cleaned_data['prix']
        entree_date = form.cleaned_data['entree_date']

        try:
            process_stock_entry(article, new_qte, new_prix, entree_date, request.user)
            messages.info(request, 'Entrée ajoutée avec succès.')
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'ajout de l\'entrée: {e}')
        return redirect('magasin:article_detail', art_id=article.art_id, slug=article.slug)


def detail_handle_sortie(request, article):
    """Handle sortie from the details page"""
    form = forms.SortieForm(request.POST)
    if form.is_valid():
        sortie_qte = form.cleaned_data['qte']
        sortie_date = form.cleaned_data['sortie_date']

        success, msg = process_stock_sortie(article, sortie_qte, sortie_date, request.user)
        if success: messages.success(request, msg)
        else: messages.error(request, msg)
        return redirect('magasin:article_detail', art_id=article.art_id, slug=article.slug)
    else:
        messages.error(request, 'Erreur lors de la sortie: données invalides.')


def detail_handle_update(request, article):
    """Edit article from the details page"""
    form = forms.ArticalForm(request.POST, instance=article)
    if form.is_valid():
        try:
            with transaction.atomic():
                form.save()
            messages.success(request, 'Article modifié avec succès.')
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification de l\'article: {e}')

        return redirect('magasin:article_detail', art_id=article.art_id, slug=article.slug)


def detail_handle_add_commande(request, article):
    """Handle add commande from the details page"""
    form = forms.CreateCommandForm(request.POST)
    if form.is_valid():
        cd = form.cleaned_data
        try:
            with transaction.atomic():
                models.Command.objects.create(
                    art_id=article, user_id=request.user,
                    command_date=cd['cmnd_date'], qte=cd['cmnd_qte']
                )
            messages.success(request, 'Commande ajoutée avec succès.')
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'ajout de la commande: {e}')
        return redirect('magasin:article_detail', art_id=article.art_id, slug=article.slug)


# 🔥 Helper function to create a Movement
def create_initial_movement(article, user, movement_date, qte, prix):
    models.Movement.objects.create(
        art_id=article,
        movement_date=movement_date,
        user_id=user,
        movement="Initial",
        qte=qte,
        prix=prix,
    )


def write_to_excel_g(fields, queryset, filename_prefix):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mouvements"

    # Write header
    for col_num, field in enumerate(fields, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = field.replace('_', ' ').capitalize()

    # Write data
    for row_num, row in enumerate(queryset.values_list(*fields), 2):
        for col_num, value in enumerate(row, 1):
            # Handle timezone-aware datetime objects
            if isinstance(value, datetime) and is_aware(value):
                value = value.replace(tzinfo=None)
                value = value.strftime('%d-%m-%Y')
            ws.cell(row=row_num, column=col_num, value=value)

    # Save with a timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{filename_prefix}_{timestamp}.xlsx"
    filepath = os.path.join('media', 'exports', filename)

    # Ensure directory exists
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)


def hijri_():
    """This functions returns a hijri date"""
    tday = date.today()
    hijri = HijriDate(tday.year, tday.month, tday.day, gr=True)
    return hijri
